import openpyxl
import xlwt
from openpyxl import load_workbook


def stat_get_win(nik):
    book = load_workbook('exem.xlsx')
    sheet = book.active
    new_data = [nik]
    search(sheet, new_data, nik)
    book.save('exem.xlsx')


def search(sheet, new_data, nik):
    flag = False
    for i in range(1, 50):
        for column in sheet['a' + str(i):'a' + str(i)]:
            for index, cell in enumerate(column):
                if cell.value == nik:
                    flag = True
                    append(sheet, i)
                elif cell.value == 1:
                    cell.value = new_data[index]
                    flag = True
                    append(sheet, i)
        if flag:
            break


def append(sheet, i):
    for row in sheet['b' + str(i):'b' + str(i)]:
        for index1, cell1 in enumerate(row):
            cell1.value = cell1.value + 1
            break
    for row in sheet['c' + str(i):'c' + str(i)]:
        for index1, cell1 in enumerate(row):
            cell1.value = cell1.value + 1
            break


def append2(sheet, i):
    for row in sheet['d' + str(i):'d' + str(i)]:
        for index1, cell1 in enumerate(row):
            cell1.value = cell1.value + 1
            break
    for row in sheet['c' + str(i):'c' + str(i)]:
        for index1, cell1 in enumerate(row):
            cell1.value = cell1.value + 1
            break


def search2(sheet, new_data, nik2):
    flag = False
    for i in range(1, 50):
        for column in sheet['a' + str(i):'a' + str(i)]:
            for index, cell in enumerate(column):
                if cell.value == nik2:
                    flag = True
                    append2(sheet, i)
                elif cell.value == 1:
                    cell.value = new_data[index]
                    flag = True
                    append2(sheet, i)
        if flag:
            break


def stat_get_loose(nik2):
    book = load_workbook('exem.xlsx')
    sheet = book.active
    new_data = [nik2]
    search2(sheet, new_data, nik2)
    book.save('exem.xlsx')


def stat_get_draw2(nik2):
    book = load_workbook('exem.xlsx')
    sheet = book.active
    new_data = [nik2]
    search3(sheet, new_data, nik2)
    book.save('exem.xlsx')


def stat_get_draw(nik):
    book = load_workbook('exem.xlsx')
    sheet = book.active
    new_data = [nik]
    search4(sheet, new_data, nik)
    book.save('exem.xlsx')


def search3(sheet, new_data, nik2):
    flag = False
    for i in range(1, 50):
        for column in sheet['a' + str(i):'a' + str(i)]:
            for index, cell in enumerate(column):
                if cell.value == nik2:
                    flag = True
                    append3(sheet, i)
                elif cell.value == 1:
                    cell.value = new_data[index]
                    flag = True
                    append3(sheet, i)
        if flag:
            break


def search4(sheet, new_data, nik):
    flag = False
    for i in range(1, 50):
        for column in sheet['a' + str(i):'a' + str(i)]:
            for index, cell in enumerate(column):
                if cell.value == nik:
                    flag = True
                    append3(sheet, i)
                elif cell.value == 1:
                    cell.value = new_data[index]
                    flag = True
                    append3(sheet, i)
        if flag:
            break


def append3(sheet, i):
    for row in sheet['c' + str(i):'c' + str(i)]:
        for index1, cell1 in enumerate(row):
            cell1.value = cell1.value + 1
            break


def stat_give(bot, chat_id, message):
    book = load_workbook('exem.xlsx')
    sheet = book.active
    user = message.from_user.username
    search4(sheet, user, bot, chat_id)
    book.save('exem.xlsx')


def search4(sheet, user, bot, chat_id):
    flag = False
    flag2 = True
    for i in range(2, 50):
        for column in sheet['a' + str(i):'a' + str(i)]:
            for index, cell in enumerate(column):
                if cell.value == user:
                    flag = True
                    flag2 = False
                    apen(sheet, i, bot, chat_id)
        if flag:
            break
    if flag2:
        bot.send_message(chat_id, 'Вы не сыграли не одной игры ещё')


def apen(sheet, i, bot, chat_id):
    for row in sheet['b' + str(i):'b' + str(i)]:
        for index1, cell1 in enumerate(row):
            win = cell1.value
            bot.send_message(chat_id, 'Побед:' + str(win))
            break
    for row in sheet['c' + str(i):'c' + str(i)]:
        for index1, cell1 in enumerate(row):
            game = cell1.value
            bot.send_message(chat_id, 'Сыграно игр:' + str(game))
            break
    for row in sheet['d' + str(i):'d' + str(i)]:
        for index1, cell1 in enumerate(row):
            loose = cell1.value
            bot.send_message(chat_id, 'Проигрышей:' + str(loose))
            break
